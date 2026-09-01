"""
File attachment handling for the AI Assurance Lab chat.

Uploaded files are processed entirely in memory and never written to disk or
persisted anywhere: they exist only for the lifetime of a single chat request
and are converted into Claude message content blocks (images) or extracted
text (PDF/XLS/CSV) before being discarded. This avoids the storage, cleanup,
and path-traversal risks that come with saving user uploads to a filesystem.
"""

import base64
import logging
import os

from pypdf import PdfReader
from openpyxl import load_workbook
import io

logger = logging.getLogger(__name__)

# Allow-list of accepted extensions -> category. Extension is checked after
# the filename is taken at face value from the client, so this is combined
# with a magic-byte sanity check per category below.
ALLOWED_EXTENSIONS = {
    'png': 'image', 'jpg': 'image', 'jpeg': 'image', 'gif': 'image', 'webp': 'image',
    'pdf': 'pdf',
    'xls': 'excel', 'xlsx': 'excel',
    'csv': 'csv',
}

IMAGE_MEDIA_TYPES = {
    'png': 'image/png', 'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
    'gif': 'image/gif', 'webp': 'image/webp',
}

MAX_FILES = 5
MAX_FILE_BYTES = 10 * 1024 * 1024  # 10 MB per file
MAX_EXTRACTED_TEXT_CHARS = 20000  # cap per-document extracted text sent to the model

IMAGE_MAGIC_BYTES = {
    'png': [b'\x89PNG\r\n\x1a\n'],
    'jpg': [b'\xff\xd8\xff'],
    'jpeg': [b'\xff\xd8\xff'],
    'gif': [b'GIF87a', b'GIF89a'],
    'webp': [b'RIFF'],  # followed by 'WEBP' at offset 8, checked separately
}


class AttachmentError(Exception):
    """Raised when an uploaded attachment is invalid or cannot be processed."""
    pass


def _get_extension(filename: str) -> str:
    if not filename or '.' not in filename:
        raise AttachmentError(f"File '{filename}' has no extension; unable to determine type")
    ext = filename.rsplit('.', 1)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        allowed = ', '.join(sorted(set(ALLOWED_EXTENSIONS.keys())))
        raise AttachmentError(f"File type '.{ext}' is not supported. Allowed types: {allowed}")
    return ext


def _verify_image_signature(ext: str, data: bytes) -> None:
    """Best-effort magic-byte check so a renamed non-image can't slip through as one."""
    if ext == 'webp':
        if not (data[:4] == b'RIFF' and data[8:12] == b'WEBP'):
            raise AttachmentError("File content does not match a valid WEBP image")
        return
    signatures = IMAGE_MAGIC_BYTES.get(ext, [])
    if signatures and not any(data.startswith(sig) for sig in signatures):
        raise AttachmentError(f"File content does not match a valid .{ext} image")


def _verify_pdf_signature(data: bytes) -> None:
    if not data.startswith(b'%PDF-'):
        raise AttachmentError("File content does not match a valid PDF")


def _extract_pdf_text(data: bytes, filename: str) -> str:
    try:
        reader = PdfReader(io.BytesIO(data))
    except Exception as e:
        raise AttachmentError(f"Could not parse PDF '{filename}': {type(e).__name__}")

    pages_text = []
    for page in reader.pages[:30]:  # cap pages read to bound processing time
        try:
            pages_text.append(page.extract_text() or '')
        except Exception:
            continue

    text = '\n'.join(pages_text).strip()
    if not text:
        text = '[No extractable text found in this PDF - it may be a scanned/image-only document]'
    if len(text) > MAX_EXTRACTED_TEXT_CHARS:
        text = text[:MAX_EXTRACTED_TEXT_CHARS] + '\n... [truncated]'
    return text


def _extract_excel_text(data: bytes, filename: str) -> str:
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise AttachmentError(f"Could not parse spreadsheet '{filename}': {type(e).__name__}")

    chunks = []
    total_chars = 0
    for sheet_name in wb.sheetnames:
        if total_chars > MAX_EXTRACTED_TEXT_CHARS:
            chunks.append('... [additional sheets truncated]')
            break
        sheet = wb[sheet_name]
        chunks.append(f"--- Sheet: {sheet_name} ---")
        row_count = 0
        for row in sheet.iter_rows(values_only=True):
            if row_count >= 500 or total_chars > MAX_EXTRACTED_TEXT_CHARS:
                chunks.append('... [truncated]')
                break
            line = ','.join('' if v is None else str(v) for v in row)
            chunks.append(line)
            total_chars += len(line)
            row_count += 1

    text = '\n'.join(chunks).strip()
    if len(text) > MAX_EXTRACTED_TEXT_CHARS:
        text = text[:MAX_EXTRACTED_TEXT_CHARS] + '\n... [truncated]'
    return text or '[Spreadsheet appears to be empty]'


def _extract_csv_text(data: bytes, filename: str) -> str:
    try:
        text = data.decode('utf-8', errors='replace')
    except Exception as e:
        raise AttachmentError(f"Could not read CSV '{filename}': {type(e).__name__}")
    if len(text) > MAX_EXTRACTED_TEXT_CHARS:
        text = text[:MAX_EXTRACTED_TEXT_CHARS] + '\n... [truncated]'
    return text


def process_uploaded_files(files) -> list:
    """
    Validate and convert a list of Werkzeug FileStorage objects into Claude
    Messages API content blocks. Images become base64 image blocks; PDFs,
    spreadsheets, and CSVs are text-extracted and returned as labeled text
    blocks so the model can reason over their contents.

    Raises AttachmentError with a user-facing message on any validation failure.
    """
    if len(files) > MAX_FILES:
        raise AttachmentError(f"Too many files attached (max {MAX_FILES})")

    content_blocks = []

    for f in files:
        filename = os.path.basename(f.filename or 'upload')
        ext = _get_extension(filename)
        category = ALLOWED_EXTENSIONS[ext]

        data = f.read()
        if not data:
            raise AttachmentError(f"File '{filename}' is empty")
        if len(data) > MAX_FILE_BYTES:
            raise AttachmentError(
                f"File '{filename}' is too large ({len(data) // 1024 // 1024} MB). "
                f"Max size is {MAX_FILE_BYTES // 1024 // 1024} MB per file."
            )

        if category == 'image':
            _verify_image_signature(ext, data)
            content_blocks.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": IMAGE_MEDIA_TYPES[ext],
                    "data": base64.b64encode(data).decode('ascii')
                }
            })
        elif category == 'pdf':
            _verify_pdf_signature(data)
            extracted = _extract_pdf_text(data, filename)
            content_blocks.append({
                "type": "text",
                "text": f"[Attached PDF: {filename}]\n{extracted}"
            })
        elif category == 'excel':
            extracted = _extract_excel_text(data, filename)
            content_blocks.append({
                "type": "text",
                "text": f"[Attached spreadsheet: {filename}]\n{extracted}"
            })
        elif category == 'csv':
            extracted = _extract_csv_text(data, filename)
            content_blocks.append({
                "type": "text",
                "text": f"[Attached CSV: {filename}]\n{extracted}"
            })

        logger.info(f"Processed attachment '{filename}' ({category}, {len(data)} bytes)")

    return content_blocks
