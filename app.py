"""
AI Assurance Lab - Flask application with user-configurable MCP connectivity.
Each student manages their own encrypted API credentials for ThousandEyes and Meraki.
"""

import os
import json
import logging
import secrets
from datetime import datetime, timedelta, timezone
from functools import wraps
from urllib.parse import urlencode, parse_qs
from urllib.request import urlopen

from flask import Flask, render_template, session, request, redirect, url_for, jsonify
from flask_cors import CORS
import requests
import boto3
import jwt
from jwt import PyJWKClient

from dynamo_db import (
    save_user_credentials,
    get_user_credentials,
    test_te_connectivity,
    test_meraki_connectivity,
    test_splunk_connectivity,
    update_connection_status,
    delete_user_credentials,
    DynamoDBError
)
from crypto import EncryptionError
from mcp_client import (
    list_mcp_tools,
    call_mcp_tool,
    MCPClientError,
    THOUSANDEYES_MCP_URL,
    MERAKI_MCP_URL
)
from attachments import process_uploaded_files, AttachmentError, MAX_FILES, MAX_FILE_BYTES

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)

_secret_key = os.getenv('SECRET_KEY')
if not _secret_key:
    # Failing fast here is deliberate: a silent hardcoded fallback would let
    # anyone who reads the (public) source code forge session cookies for
    # any account, including proctors, if SECRET_KEY were ever unset.
    raise RuntimeError(
        "SECRET_KEY environment variable is required and must not be empty. "
        "Refusing to start with an insecure default."
    )
app.secret_key = _secret_key

# Cap total request size well above MAX_FILES * MAX_FILE_BYTES to leave room
# for form fields/history, while still bounding worst-case memory/DoS exposure.
app.config['MAX_CONTENT_LENGTH'] = (MAX_FILES * MAX_FILE_BYTES) + (2 * 1024 * 1024)

# Session cookie hardening: the app is HTTPS-only (Nginx redirects http->https),
# cookies never need to leave this origin, and no page needs JS access to the
# cookie - so lock all three flags down explicitly rather than relying on
# framework defaults, which vary by Flask version and aren't guaranteed Secure.
app.config['SESSION_COOKIE_SECURE'] = True
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# This app isn't a public cross-origin API - every request comes from this
# app's own pages. Restrict CORS to the app's own origin instead of the
# flask_cors default of allowing any origin.
CORS(app, origins=[os.getenv('APP_URL', '')], supports_credentials=True)

HTML_OUTPUT_SYSTEM_PROMPT = (
    "When the user asks you to generate a report, document, dashboard, or any kind of "
    "downloadable output, produce a complete, valid, self-contained HTML document "
    "(inline <style>, no external resources or network requests) inside a single "
    "```html fenced code block. The user's chat interface will automatically offer "
    "a download button for any ```html code block you return, so prefer this format "
    "whenever the user wants something they can save or share, rather than only "
    "describing it in prose."
)


def _build_system_prompt(meraki_org_id=None):
    """Build the system prompt, including the real current date/time.

    Without this, the model has no way to know "now" and will guess a date from
    its training data when reasoning about relative time windows (e.g. "the last
    2 hours"), which can silently produce the wrong absolute timestamps for tool
    calls and mislead users in the response text.
    """
    now_utc = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')
    prompt = (
        f"The current date and time is {now_utc}. Use this as \"now\" whenever you "
        "need to reason about or compute relative time windows (e.g. \"the last 2 "
        "hours\", \"today\", \"this week\") for tool calls or in your response. "
        "Do not guess or infer the date from any other source.\n\n"
        + HTML_OUTPUT_SYSTEM_PROMPT
    )
    if meraki_org_id:
        prompt += (
            f"\n\nThe user's Meraki Organization ID is {meraki_org_id}. Use it "
            "automatically for any Meraki tool call that requires an organization "
            "ID (e.g. via execute_api parameters) - never ask the user for it."
        )
    return prompt

# AWS and Cognito configuration
COGNITO_DOMAIN = os.getenv('COGNITO_DOMAIN')
COGNITO_CLIENT_ID = os.getenv('COGNITO_CLIENT_ID')
COGNITO_CLIENT_SECRET = os.getenv('COGNITO_CLIENT_SECRET')
COGNITO_REGION = os.getenv('COGNITO_REGION', 'us-east-1')
COGNITO_USER_POOL_ID = os.getenv('COGNITO_USER_POOL_ID')
APP_URL = os.getenv('APP_URL', 'http://localhost:5000')
BEDROCK_REGION = os.getenv('BEDROCK_REGION', 'us-east-1')

# Initialize Bedrock client
bedrock_client = boto3.client('bedrock-runtime', region_name=BEDROCK_REGION)


def login_required(f):
    """Decorator to require Cognito login."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_email' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


# Initialize Cognito client for user management
cognito_client = boto3.client('cognito-idp', region_name=COGNITO_REGION)

COGNITO_ISSUER = f'https://cognito-idp.{COGNITO_REGION}.amazonaws.com/{COGNITO_USER_POOL_ID}'
# PyJWKClient fetches and caches Cognito's public signing keys (JWKS) and
# picks the right one per token's `kid` header - this is what lets us verify
# the ID token's RS256 signature instead of trusting it blindly.
_jwks_client = PyJWKClient(f'{COGNITO_ISSUER}/.well-known/jwks.json')


# ============================================================================
# Authentication Routes
# ============================================================================

@app.route('/login')
def login():
    """Redirect to Cognito login."""
    # CSRF protection for the OAuth flow itself ("login CSRF"): without a
    # state value tied to this browser's session, an attacker could get
    # their own valid `code` from Cognito and trick a victim into visiting
    # /auth/callback?code=<attacker's code>, silently signing the victim in
    # as the attacker's identity. Binding a random state to the session and
    # checking it on callback prevents that.
    state = secrets.token_urlsafe(24)
    session['oauth_state'] = state
    return redirect(
        f'https://{COGNITO_DOMAIN}/oauth2/authorize?'
        f'client_id={COGNITO_CLIENT_ID}&'
        f'response_type=code&'
        f'redirect_uri={APP_URL}/auth/callback&'
        f'scope=email+openid+profile&'
        f'state={state}'
    )


@app.route('/auth/callback')
def auth_callback():
    """Handle Cognito callback after login."""
    code = request.args.get('code')
    returned_state = request.args.get('state')
    expected_state = session.pop('oauth_state', None)

    if not code:
        logger.warning("Auth callback received without code")
        return jsonify({'error': 'No authorization code received'}), 400

    # Reject if there's no state to check against, or it doesn't match what
    # this browser's session was given in /login. See the comment in
    # login() for why this matters (CSRF / login fixation protection).
    if not expected_state or not returned_state or returned_state != expected_state:
        logger.warning("Auth callback rejected: missing or mismatched OAuth state")
        return jsonify({'error': 'Invalid or expired login attempt. Please try logging in again.'}), 400

    try:
        # Exchange code for tokens
        token_url = f'https://{COGNITO_DOMAIN}/oauth2/token'
        token_data = {
            'grant_type': 'authorization_code',
            'client_id': COGNITO_CLIENT_ID,
            'client_secret': COGNITO_CLIENT_SECRET,
            'code': code,
            'redirect_uri': f'{APP_URL}/auth/callback'
        }
        
        token_response = requests.post(token_url, data=token_data, timeout=10)
        token_response.raise_for_status()
        
        tokens = token_response.json()
        id_token = tokens.get('id_token')
        
        if not id_token:
            logger.error("No ID token in Cognito response")
            return jsonify({'error': 'Failed to obtain ID token'}), 400
        
        # Verify the ID token properly: signature (via Cognito's published
        # JWKS), issuer, audience, and expiry. Even though this token came
        # from a direct server-to-server call to Cognito above (not
        # something a client could inject), verifying it fully is cheap
        # defense-in-depth and is what every Cognito integration guide
        # requires - a previous version of this code skipped signature
        # verification entirely, which is not something to carry forward.
        try:
            signing_key = _jwks_client.get_signing_key_from_jwt(id_token)
            decoded = jwt.decode(
                id_token,
                signing_key.key,
                algorithms=["RS256"],
                audience=COGNITO_CLIENT_ID,
                issuer=COGNITO_ISSUER,
            )
            user_email = decoded.get('email')
            
            if not user_email:
                logger.error("No email in ID token")
                return jsonify({'error': 'Email not found in token'}), 400
            
            # Store in session
            session['user_email'] = user_email
            session['access_token'] = tokens.get('access_token')
            session.permanent = True
            app.permanent_session_lifetime = timedelta(days=7)
            
            logger.info(f"User logged in: {user_email}")
            
            return redirect(url_for('lab'))
        
        except jwt.InvalidTokenError as e:
            logger.error(f"ID token failed verification: {e}")
            return jsonify({'error': 'Failed to verify identity token'}), 400
    
    except requests.RequestException as e:
        logger.error(f"Cognito token exchange failed: {e}")
        return jsonify({'error': 'Authentication failed'}), 500


@app.route('/logout')
def logout():
    """Clear session and redirect to Cognito logout."""
    session.clear()
    return redirect(
        f'https://{COGNITO_DOMAIN}/logout?'
        f'client_id={COGNITO_CLIENT_ID}&'
        f'logout_uri={APP_URL}'
    )


# ============================================================================
# Credential Management Routes
# ============================================================================

@app.route('/credentials')
@login_required
def credentials_page():
    """Display MCP credential management UI."""
    return render_template('credentials.html', email=session.get('user_email'))


@app.route('/api/credentials', methods=['GET'])
@login_required
def get_credentials_status():
    """
    Get user's MCP credential status (no plaintext tokens).
    Returns configuration status and connection state.
    """
    try:
        email = session.get('user_email')
        credentials = get_user_credentials(email)
        
        return jsonify({
            "te_configured": bool(credentials.get('te_token')),
            "te_connected": credentials.get('te_connected', False),
            "meraki_configured": bool(credentials.get('meraki_token')),
            "meraki_connected": credentials.get('meraki_connected', False),
            "meraki_org_id": credentials.get('meraki_org_id'),
            "splunk_configured": bool(credentials.get('splunk_url')),
            "splunk_connected": credentials.get('splunk_connected', False),
            "splunk_url": credentials.get('splunk_url'),
            "last_updated": credentials.get('updated_at')
        })
    
    except DynamoDBError as e:
        logger.error(f"DynamoDB error for {email}: {e}")
        return jsonify({'error': 'Failed to retrieve credentials'}), 500
    except Exception as e:
        logger.error(f"Unexpected error in get_credentials_status: {e}")
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/api/credentials/add', methods=['POST'])
@login_required
def add_credential():
    """
    Add or update a user's MCP credential.
    Validates token format, encrypts, and stores in DynamoDB.
    """
    try:
        email = session.get('user_email')
        data = request.json
        service = data.get('service', '').lower()
        token = (data.get('token') or '').strip()
        url = (data.get('url') or '').strip()
        org_id = (data.get('org_id') or '').strip()
        
        if service not in ['thousandeyes', 'meraki', 'splunk']:
            logger.warning(f"add_credential rejected for {email}: unknown service '{service}'")
            return jsonify({'error': 'Unknown service'}), 400
        
        if service == 'splunk':
            # Splunk's MCP server URL varies per student (local laptop via a
            # tunnel, or a facilitator-hosted shared server). The token/API
            # key is optional - some self-hosted servers don't require one.
            if not url:
                logger.warning(f"add_credential rejected for {email}: splunk url missing")
                return jsonify({'error': 'Splunk MCP server URL is required'}), 400
            if not url.startswith('http://') and not url.startswith('https://'):
                logger.warning(f"add_credential rejected for {email}: splunk url '{url}' missing http(s) scheme")
                return jsonify({'error': 'Splunk MCP server URL must start with http:// or https://'}), 400
            save_user_credentials(email, splunk_url=url, splunk_token=token or None)
            logger.info(f"Saved splunk credential for {email}")
            return jsonify({'status': 'success', 'message': 'Splunk credential saved'})
        
        # Meraki also takes an Organization ID (not a secret), which the
        # Meraki MCP server's tools require for almost every call - there's
        # no reliable "list my organizations" step, so we store it once here
        # instead of asking the student to paste it into every chat message.
        if service == 'meraki' and org_id and not token:
            save_user_credentials(email, meraki_org_id=org_id)
            logger.info(f"Saved meraki org_id for {email}")
            return jsonify({'status': 'success', 'message': 'Meraki organization ID saved'})
        
        if not token:
            logger.warning(f"add_credential rejected for {email}: {service} token missing")
            return jsonify({'error': 'token is required'}), 400
        
        # Validate token format
        if service == 'thousandeyes':
            # ThousandEyes tokens are typically 32+ chars
            if len(token) < 10:
                logger.warning(f"add_credential rejected for {email}: thousandeyes token too short (len={len(token)})")
                return jsonify({'error': 'ThousandEyes token appears too short'}), 400
        
        elif service == 'meraki':
            # Meraki tokens are typically 32+ chars
            if len(token) < 20:
                logger.warning(f"add_credential rejected for {email}: meraki token too short (len={len(token)})")
                return jsonify({'error': 'Meraki token appears too short'}), 400
        
        # Save credentials
        if service == 'thousandeyes':
            save_user_credentials(email, te_token=token)
        else:
            save_user_credentials(email, meraki_token=token, meraki_org_id=org_id or None)
        
        logger.info(f"Saved {service} credential for {email}")
        
        return jsonify({
            'status': 'success',
            'message': f'{service.capitalize()} credential saved'
        })
    
    except EncryptionError as e:
        logger.error(f"Encryption error for {email}: {e}")
        return jsonify({'error': 'Failed to encrypt credential'}), 500
    except DynamoDBError as e:
        logger.error(f"DynamoDB error for {email}: {e}")
        return jsonify({'error': 'Failed to save credential'}), 500
    except Exception as e:
        logger.error(f"Error in add_credential: {e}")
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/api/credentials/test', methods=['POST'])
@login_required
def test_credential():
    """
    Test if a credential is valid by making a test API call.
    Updates connection status in DynamoDB.
    """
    try:
        email = session.get('user_email')
        data = request.json
        service = data.get('service', '').lower()
        
        if service not in ['thousandeyes', 'meraki', 'splunk']:
            return jsonify({'error': 'Unknown service'}), 400
        
        # Test connectivity
        if service == 'thousandeyes':
            result = test_te_connectivity(email)
        elif service == 'meraki':
            result = test_meraki_connectivity(email)
        else:
            result = test_splunk_connectivity(email)
        
        # Update connection status
        is_valid = result.get('valid', False)
        update_connection_status(email, service, is_valid)
        
        if is_valid:
            logger.info(f"Connectivity test passed for {service} on {email}")
            return jsonify({
                'status': 'success',
                'connected': True,
                'message': f'{service.capitalize()} credential is valid'
            })
        else:
            error_msg = result.get('error', f'{service} credential test failed')
            logger.warning(f"Connectivity test failed for {service} on {email}: {error_msg}")
            return jsonify({
                'status': 'error',
                'connected': False,
                'message': error_msg
            }), 400
    
    except DynamoDBError as e:
        logger.error(f"DynamoDB error during credential test: {e}")
        return jsonify({'error': 'Failed to test credential'}), 500
    except Exception as e:
        logger.error(f"Error in test_credential: {e}")
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/api/credentials/delete', methods=['POST'])
@login_required
def delete_credential():
    """
    Revoke a user's MCP credential for a service.
    """
    try:
        email = session.get('user_email')
        data = request.json
        service = data.get('service', '').lower()
        
        if service not in ['thousandeyes', 'meraki', 'splunk']:
            return jsonify({'error': 'Unknown service'}), 400
        
        delete_user_credentials(email, service)
        
        logger.info(f"Deleted {service} credential for {email}")
        
        return jsonify({
            'status': 'success',
            'message': f'{service.capitalize()} credential deleted'
        })
    
    except DynamoDBError as e:
        logger.error(f"DynamoDB error deleting credential: {e}")
        return jsonify({'error': 'Failed to delete credential'}), 500
    except Exception as e:
        logger.error(f"Error in delete_credential: {e}")
        return jsonify({'error': 'Internal server error'}), 500


# ============================================================================
# Chat Route (with user credentials)
# ============================================================================

@app.route('/lab')
@login_required
def lab():
    """Display the main lab chat interface."""
    email = session.get('user_email')
    return render_template('lab.html', email=email)


@app.route('/guide')
@login_required
def guide():
    """Display the lab guide (rendered standalone, embedded via iframe in the lab sidebar)."""
    return render_template('guide.html')


@app.route('/api/chat', methods=['POST'])
@login_required
def chat():
    """
    Chat with Claude using user's stored MCP credentials.
    Retrieves per-user tokens and only exposes tools for configured services.
    """
    try:
        email = session.get('user_email')

        # Support both plain JSON (no attachments) and multipart/form-data
        # (attachments present) requests from the frontend.
        if request.content_type and 'multipart/form-data' in request.content_type:
            user_message = request.form.get('message', '')
            try:
                conversation_history = json.loads(request.form.get('history', '[]'))
            except (TypeError, ValueError):
                conversation_history = []
            uploaded_files = request.files.getlist('files')
        else:
            data = request.json or {}
            user_message = data.get('message', '')
            conversation_history = data.get('history', [])
            uploaded_files = []

        if not user_message and not uploaded_files:
            return jsonify({'error': 'Message is required'}), 400

        # Process attachments in-memory only; nothing is written to disk.
        attachment_blocks = []
        if uploaded_files:
            try:
                attachment_blocks = process_uploaded_files(uploaded_files)
            except AttachmentError as e:
                return jsonify({'error': str(e)}), 400

        # Get user's stored credentials
        try:
            credentials = get_user_credentials(email)
        except DynamoDBError as e:
            logger.error(f"Failed to retrieve credentials for {email}: {e}")
            return jsonify({'error': 'Failed to retrieve your credentials'}), 500
        
        te_token = credentials.get('te_token')
        meraki_token = credentials.get('meraki_token')
        meraki_org_id = credentials.get('meraki_org_id')
        splunk_url = credentials.get('splunk_url')
        splunk_token = credentials.get('splunk_token')
        
        # Discover live tools from the user's own MCP servers (ThousandEyes and
        # Meraki host their own MCP servers; we call them with the user's token
        # rather than proxying REST calls ourselves).
        available_tools = []
        tool_routing = {}  # tool_name -> (mcp_url, token, require_token)
        
        if te_token:
            try:
                te_tools = list_mcp_tools(THOUSANDEYES_MCP_URL, te_token)
                for tool in te_tools:
                    available_tools.append(tool)
                    tool_routing[tool['name']] = (THOUSANDEYES_MCP_URL, te_token, True)
            except MCPClientError as e:
                logger.warning(f"Failed to list ThousandEyes MCP tools for {email}: {e}")
        
        if meraki_token:
            try:
                meraki_tools = list_mcp_tools(MERAKI_MCP_URL, meraki_token)
                for tool in meraki_tools:
                    available_tools.append(tool)
                    tool_routing[tool['name']] = (MERAKI_MCP_URL, meraki_token, True)
            except MCPClientError as e:
                logger.warning(f"Failed to list Meraki MCP tools for {email}: {e}")
        
        # Splunk's MCP server URL is student-configured (their own laptop via
        # a tunnel, or a facilitator-hosted shared server), and some setups
        # don't require an auth token, so we don't gate on one being present.
        if splunk_url:
            try:
                splunk_tools = list_mcp_tools(splunk_url, splunk_token, require_token=False)
                for tool in splunk_tools:
                    available_tools.append(tool)
                    tool_routing[tool['name']] = (splunk_url, splunk_token, False)
            except MCPClientError as e:
                logger.warning(f"Failed to list Splunk MCP tools for {email}: {e}")
        
        # Build messages for Claude. If there are attachments, the user turn
        # becomes a list of content blocks (images/extracted text + the
        # user's own text) instead of a plain string.
        if attachment_blocks:
            user_content = list(attachment_blocks)
            if user_message:
                user_content.append({"type": "text", "text": user_message})
        else:
            user_content = user_message

        messages = conversation_history + [
            {"role": "user", "content": user_content}
        ]

        # Agentic loop: Claude may need multiple rounds of tool calls before it
        # has enough information to answer (e.g. Meraki's MCP server exposes a
        # generic semantic_search + execute_api pair that often requires a
        # search call followed by an execute call). Cap iterations to avoid
        # runaway loops while still allowing multi-step tool use.
        MAX_TOOL_ITERATIONS = 6
        total_tools_used = 0
        content = []
        system_prompt = _build_system_prompt(meraki_org_id)

        for iteration in range(MAX_TOOL_ITERATIONS):
            request_body = {
                "anthropic_version": "bedrock-2023-05-31",
                "max_tokens": 4096,
                "system": system_prompt,
                "messages": messages
            }
            if available_tools:
                request_body["tools"] = available_tools

            try:
                response = bedrock_client.invoke_model(
                    modelId='us.anthropic.claude-sonnet-4-5-20250929-v1:0',
                    body=json.dumps(request_body)
                )
                result = json.loads(response['body'].read())
                content = result.get('content', [])
            except Exception as e:
                logger.error(f"Bedrock invocation failed: {e}")
                return jsonify({'error': 'Failed to invoke AI model'}), 500

            tool_calls = [c for c in content if c.get('type') == 'tool_use']

            if not tool_calls:
                # Claude is done calling tools; this is the final answer.
                break

            total_tools_used += len(tool_calls)
            tool_results = []

            for tool_call in tool_calls:
                tool_name = tool_call.get('name', '')
                tool_input = tool_call.get('input', {})
                tool_use_id = tool_call.get('id', '')

                # Route to the MCP server (ThousandEyes or Meraki) that advertised this tool
                try:
                    routing = tool_routing.get(tool_name)
                    if not routing:
                        tool_result = {"error": f"Unknown tool: {tool_name}"}
                    else:
                        mcp_url, mcp_token, require_token = routing
                        tool_result = call_mcp_tool(mcp_url, mcp_token, tool_name, tool_input, require_token=require_token)
                except MCPClientError as e:
                    logger.error(f"Tool execution error for {tool_name}: {e}")
                    tool_result = {"error": f"Tool execution failed: {str(e)}"}
                except Exception as e:
                    logger.error(f"Tool execution error for {tool_name}: {e}")
                    tool_result = {"error": f"Tool execution failed: {str(e)}"}

                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": tool_use_id,
                    "content": json.dumps(tool_result) if not isinstance(tool_result, str) else tool_result
                })

            # Feed the tool results back and let Claude continue (may call more tools)
            messages.append({"role": "assistant", "content": content})
            messages.append({"role": "user", "content": tool_results})
        else:
            logger.warning(f"Hit max tool iterations ({MAX_TOOL_ITERATIONS}) for {email}")

        assistant_message = next(
            (c.get('text', '') for c in content if c.get('type') == 'text'),
            'The assistant used tools but did not return a final text response. Please try rephrasing your question.'
        )

        return jsonify({
            "response": assistant_message,
            "tools_used": total_tools_used,
            "te_available": bool(te_token),
            "meraki_available": bool(meraki_token),
            "splunk_available": bool(splunk_url)
        })
    
    except Exception as e:
        logger.error(f"Error in chat endpoint: {e}")
        return jsonify({'error': 'Internal server error'}), 500


# ============================================================================
# Student Management Routes (Proctor Portal)
# ============================================================================

@app.route('/admin/students')
@login_required
def admin_students():
    """Proctor portal for managing students."""
    # Only allow if user is a proctor (hardcoded for now, could use DynamoDB)
    user_email = session.get('user_email', '')
    if not _is_proctor(user_email):
        return jsonify({'error': 'Access denied - proctor access required'}), 403
    
    return render_template('admin_students.html', email=user_email)


# Hardcoded super-admin — always a proctor, can never be deleted or demoted
# via the web UI, regardless of what PROCTOR_EMAILS contains. This is a
# deliberate backstop so a misconfigured/emptied PROCTOR_EMAILS list can't
# lock everyone (including the account owner) out of admin access.
SUPER_ADMIN_EMAIL = 'sceddy@cisco.com'


def _get_proctor_emails() -> list:
    """Current proctor emails, always including the super admin."""
    raw = os.getenv('PROCTOR_EMAILS', '')
    emails = {e.strip().lower() for e in raw.split(',') if e.strip()}
    emails.add(SUPER_ADMIN_EMAIL.lower())
    return sorted(emails)


def _is_proctor(user_email: str) -> bool:
    return (user_email or '').strip().lower() in _get_proctor_emails()


def _persist_proctor_emails(emails: list) -> None:
    """Write PROCTOR_EMAILS back to .env and update the current process's
    environment so this worker sees the change immediately. Other Gunicorn
    workers won't see it until Flask is restarted, so callers should trigger
    a restart after calling this (see /api/admin/proctors/add and /delete).
    """
    # Never persist the super admin into the file-based list; it's implicit.
    emails = [e for e in emails if e.lower() != SUPER_ADMIN_EMAIL.lower()]
    value = ','.join(emails)

    env_file = '.env'
    env_content = {}
    if os.path.exists(env_file):
        with open(env_file, 'r') as f:
            for line in f:
                if '=' in line and not line.startswith('#'):
                    key, val = line.strip().split('=', 1)
                    env_content[key] = val

    env_content['PROCTOR_EMAILS'] = value
    with open(env_file, 'w') as f:
        for key, val in env_content.items():
            f.write(f"{key}={val}\n")

    os.environ['PROCTOR_EMAILS'] = value


def _restart_flask_service() -> None:
    """Best-effort restart so all Gunicorn workers pick up an env change."""
    import subprocess
    try:
        subprocess.run(['/usr/bin/sudo', '/usr/bin/systemctl', 'restart', 'flask-app'], timeout=10)
    except Exception as e:
        logger.warning(f"Could not restart flask-app service: {e}")


def _create_cognito_student(email: str, first_name: str = '', last_name: str = '') -> None:
    """
    Create a single student user in Cognito and let Cognito email the
    temporary password via SES (do NOT suppress the message).

    Why: accounts land in FORCE_CHANGE_PASSWORD after admin_create_user, and
    Cognito refuses to run the "Forgot password" flow for accounts in that
    state ("User password cannot be reset in the current state"). If the
    invite email is suppressed, the user has no way to ever learn their
    temporary password and is permanently locked out until an admin manually
    resets it. So the invite email is the only path in for a brand new
    account - it must be sent.

    Raises the underlying boto3 exception on failure so callers can classify it
    (e.g. UsernameExistsException vs. other errors).
    """
    cognito_client.admin_create_user(
        UserPoolId=os.getenv('COGNITO_USER_POOL_ID'),
        Username=email,
        UserAttributes=[
            {'Name': 'email', 'Value': email},
            {'Name': 'email_verified', 'Value': 'true'},
            {'Name': 'given_name', 'Value': first_name},
            {'Name': 'family_name', 'Value': last_name}
        ],
        DesiredDeliveryMediums=['EMAIL']
    )


def _parse_students_csv(file) -> list:
    import io
    import csv
    stream = io.StringIO(file.stream.read().decode('UTF-8'))
    reader = csv.DictReader(stream)
    return [
        {
            'email': (row.get('email') or '').strip(),
            'first_name': (row.get('first_name') or '').strip(),
            'last_name': (row.get('last_name') or '').strip(),
        }
        for row in reader
    ]


def _parse_students_excel(file) -> list:
    from openpyxl import load_workbook
    wb = load_workbook(file.stream, read_only=True, data_only=True)
    sheet = wb[wb.sheetnames[0]]

    rows_iter = sheet.iter_rows(values_only=True)
    header = [str(h).strip().lower() if h else '' for h in next(rows_iter, [])]

    def col_index(*names):
        for name in names:
            if name in header:
                return header.index(name)
        return None

    email_idx = col_index('email', 'e-mail', 'email address')
    first_idx = col_index('first_name', 'first name', 'firstname')
    last_idx = col_index('last_name', 'last name', 'lastname')

    if email_idx is None:
        raise ValueError("Spreadsheet must have an 'email' column header in the first row")

    students = []
    for row in rows_iter:
        if row is None or email_idx >= len(row):
            continue
        email = str(row[email_idx] or '').strip()
        if not email:
            continue
        first_name = str(row[first_idx]).strip() if first_idx is not None and first_idx < len(row) and row[first_idx] else ''
        last_name = str(row[last_idx]).strip() if last_idx is not None and last_idx < len(row) and row[last_idx] else ''
        students.append({'email': email, 'first_name': first_name, 'last_name': last_name})

    return students


@app.route('/api/admin/students/upload', methods=['POST'])
@login_required
def upload_students_csv():
    """Upload and create students from a CSV or Excel (.xlsx/.xls) file."""
    user_email = session.get('user_email', '')
    if not _is_proctor(user_email):
        return jsonify({'error': 'Access denied'}), 403

    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    filename_lower = file.filename.lower()
    try:
        if filename_lower.endswith('.csv'):
            students = _parse_students_csv(file)
        elif filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls'):
            students = _parse_students_excel(file)
        else:
            return jsonify({'error': 'File must be CSV or Excel (.xlsx/.xls) format'}), 400
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        logger.error(f"Error parsing student file: {str(e)}")
        return jsonify({'error': f'Could not read file: {type(e).__name__}'}), 400

    created = 0
    failed = 0
    errors = []

    for student in students:
        email = student['email']
        if not email:
            failed += 1
            errors.append('Empty email in row')
            continue

        try:
            _create_cognito_student(email, student['first_name'], student['last_name'])
            created += 1
        except cognito_client.exceptions.UsernameExistsException:
            failed += 1
            errors.append(f'{email}: User already exists')
        except Exception as e:
            failed += 1
            errors.append(f'{email}: {str(e)}')

    return jsonify({
        'status': 'success',
        'created': created,
        'failed': failed,
        'errors': errors[:10]  # Return first 10 errors
    })


@app.route('/api/admin/students/add', methods=['POST'])
@login_required
def add_single_student():
    """Create a single student account from the proctor portal form."""
    user_email = session.get('user_email', '')
    if not _is_proctor(user_email):
        return jsonify({'error': 'Access denied'}), 403

    data = request.json or {}
    email = (data.get('email') or '').strip()
    first_name = (data.get('first_name') or '').strip()
    last_name = (data.get('last_name') or '').strip()

    if not email or '@' not in email:
        return jsonify({'error': 'A valid email address is required'}), 400

    try:
        _create_cognito_student(email, first_name, last_name)
        return jsonify({'status': 'success', 'email': email})
    except cognito_client.exceptions.UsernameExistsException:
        return jsonify({'error': f'{email} already exists'}), 409
    except Exception as e:
        logger.error(f"Error creating student {email}: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/students/list', methods=['GET'])
@login_required
def list_students():
    """List all students in Cognito."""
    user_email = session.get('user_email', '')
    if not _is_proctor(user_email):
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        users = []
        pagination_token = None
        while True:
            kwargs = {'UserPoolId': os.getenv('COGNITO_USER_POOL_ID'), 'Limit': 60}
            if pagination_token:
                kwargs['PaginationToken'] = pagination_token
            response = cognito_client.list_users(**kwargs)
            users.extend(response.get('Users', []))
            pagination_token = response.get('PaginationToken')
            if not pagination_token:
                break

        students = []
        for user in users:
            email = next((attr['Value'] for attr in user['Attributes'] if attr['Name'] == 'email'), '')
            first_name = next((attr['Value'] for attr in user['Attributes'] if attr['Name'] == 'given_name'), '')
            last_name = next((attr['Value'] for attr in user['Attributes'] if attr['Name'] == 'family_name'), '')

            # Best-effort: a missing/broken credentials row shouldn't break the
            # whole list, just show everything as not-connected for that row.
            try:
                creds = get_user_credentials(email) if email else {}
            except Exception as e:
                logger.warning(f"Could not load credential status for {email}: {e}")
                creds = {}

            students.append({
                'email': email,
                'first_name': first_name,
                'last_name': last_name,
                'created': str(user['UserCreateDate']),
                'status': user['UserStatus'],
                'te_connected': bool(creds.get('te_connected')),
                'meraki_connected': bool(creds.get('meraki_connected')),
                'splunk_connected': bool(creds.get('splunk_connected'))
            })
        
        return jsonify({
            'status': 'success',
            'count': len(students),
            'students': students
        })
    
    except Exception as e:
        logger.error(f"Error listing students: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/students/delete-all', methods=['POST'])
@login_required
def delete_all_students():
    """Delete all students from Cognito. Proctors (and the super admin) are
    never deleted by this — they're accounts that persist across cohorts."""
    user_email = session.get('user_email', '')
    if not _is_proctor(user_email):
        return jsonify({'error': 'Access denied'}), 403

    proctor_emails = set(_get_proctor_emails())

    try:
        users = []
        pagination_token = None
        while True:
            kwargs = {'UserPoolId': os.getenv('COGNITO_USER_POOL_ID'), 'Limit': 60}
            if pagination_token:
                kwargs['PaginationToken'] = pagination_token
            response = cognito_client.list_users(**kwargs)
            users.extend(response.get('Users', []))
            pagination_token = response.get('PaginationToken')
            if not pagination_token:
                break

        deleted = 0
        skipped = 0
        for user in users:
            email = next((attr['Value'] for attr in user['Attributes'] if attr['Name'] == 'email'), '')
            if email.strip().lower() in proctor_emails:
                skipped += 1
                continue
            try:
                cognito_client.admin_delete_user(
                    UserPoolId=os.getenv('COGNITO_USER_POOL_ID'),
                    Username=user['Username']
                )
                deleted += 1
            except Exception as e:
                logger.warning(f"Failed to delete user {user['Username']}: {str(e)}")
        
        return jsonify({
            'status': 'success',
            'deleted': deleted,
            'skipped_proctors': skipped
        })
    
    except Exception as e:
        logger.error(f"Error deleting students: {str(e)}")
        return jsonify({'error': str(e)}), 500


# ============================================================================
# Proctor Management Routes (proctors only)
# ============================================================================

@app.route('/api/admin/proctors/list', methods=['GET'])
@login_required
def list_proctors():
    """List current proctor emails."""
    user_email = session.get('user_email', '')
    if not _is_proctor(user_email):
        return jsonify({'error': 'Access denied'}), 403

    proctors = [
        {'email': email, 'is_super_admin': email.lower() == SUPER_ADMIN_EMAIL.lower()}
        for email in _get_proctor_emails()
    ]
    return jsonify({'status': 'success', 'proctors': proctors})


@app.route('/api/admin/proctors/add', methods=['POST'])
@login_required
def add_proctor():
    """Promote/create a proctor account. Only existing proctors can do this.

    Creates a Cognito login for the new proctor (if one doesn't already
    exist) and adds their email to the protected PROCTOR_EMAILS list, then
    restarts Flask so every worker sees the updated proctor list.
    """
    user_email = session.get('user_email', '')
    if not _is_proctor(user_email):
        return jsonify({'error': 'Access denied'}), 403

    data = request.json or {}
    new_email = (data.get('email') or '').strip()
    first_name = (data.get('first_name') or '').strip()
    last_name = (data.get('last_name') or '').strip()

    if not new_email or '@' not in new_email:
        return jsonify({'error': 'A valid email address is required'}), 400

    if new_email.strip().lower() == SUPER_ADMIN_EMAIL.lower():
        return jsonify({'error': f'{SUPER_ADMIN_EMAIL} is already the permanent super admin'}), 400

    # Create their Cognito login if they don't already have one (e.g.
    # promoting an existing student just needs the PROCTOR_EMAILS update).
    try:
        _create_cognito_student(new_email, first_name, last_name)
    except cognito_client.exceptions.UsernameExistsException:
        pass
    except Exception as e:
        logger.error(f"Error creating proctor account {new_email}: {str(e)}")
        return jsonify({'error': str(e)}), 500

    current = set(_get_proctor_emails())
    current.add(new_email.strip().lower())
    _persist_proctor_emails(sorted(current))

    logger.info(f"Proctor {new_email} added by {user_email}")
    _restart_flask_service()

    return jsonify({
        'status': 'success',
        'message': f'{new_email} is now a proctor. Flask is restarting to apply it everywhere.'
    })


@app.route('/api/admin/proctors/delete', methods=['POST'])
@login_required
def delete_proctor():
    """Remove proctor status and delete the account. The super admin can
    never be removed this way."""
    user_email = session.get('user_email', '')
    if not _is_proctor(user_email):
        return jsonify({'error': 'Access denied'}), 403

    data = request.json or {}
    target_email = (data.get('email') or '').strip()

    if not target_email:
        return jsonify({'error': 'email is required'}), 400

    if target_email.strip().lower() == SUPER_ADMIN_EMAIL.lower():
        return jsonify({'error': f'{SUPER_ADMIN_EMAIL} can never be removed'}), 403

    current = set(_get_proctor_emails())
    current.discard(target_email.strip().lower())
    _persist_proctor_emails(sorted(current))

    try:
        cognito_client.admin_delete_user(
            UserPoolId=os.getenv('COGNITO_USER_POOL_ID'),
            Username=target_email
        )
    except Exception as e:
        logger.warning(f"Could not delete Cognito user for removed proctor {target_email}: {e}")

    logger.info(f"Proctor {target_email} removed by {user_email}")
    _restart_flask_service()

    return jsonify({
        'status': 'success',
        'message': f'{target_email} is no longer a proctor. Flask is restarting to apply it everywhere.'
    })

# ============================================================================
# Settings/Administration Routes
# ============================================================================

@app.route('/admin/settings')
@login_required
def admin_settings():
    """Settings and administration page for proctors."""
    user_email = session.get('user_email', '')
    if not _is_proctor(user_email):
        return jsonify({'error': 'Access denied - proctor access required'}), 403
    
    return render_template('admin_settings.html', email=user_email)


@app.route('/api/admin/settings/config', methods=['GET'])
@login_required
def get_settings():
    """Get current configuration (sensitive values redacted)."""
    user_email = session.get('user_email', '')
    if not _is_proctor(user_email):
        return jsonify({'error': 'Access denied'}), 403
    
    return jsonify({
        'cognito_domain': os.getenv('COGNITO_DOMAIN', ''),
        'cognito_client_id': os.getenv('COGNITO_CLIENT_ID', ''),
        'cognito_region': os.getenv('COGNITO_REGION', 'us-east-1'),
        'app_url': os.getenv('APP_URL', ''),
        'bedrock_region': os.getenv('BEDROCK_REGION', 'us-east-1'),
        'dynamodb_region': os.getenv('DYNAMODB_REGION', 'us-east-1'),
        'dynamodb_table': os.getenv('DYNAMODB_TABLE', ''),
        'proctor_emails': os.getenv('PROCTOR_EMAILS', ''),
        'encryption_key_set': bool(os.getenv('ENCRYPTION_KEY')),
        'flask_env': os.getenv('FLASK_ENV', 'development'),
        'has_cognito_secret': bool(os.getenv('COGNITO_CLIENT_SECRET'))
    })


@app.route('/api/admin/settings/update', methods=['POST'])
@login_required
def update_settings():
    """Update configuration settings."""
    user_email = session.get('user_email', '')
    if not _is_proctor(user_email):
        return jsonify({'error': 'Access denied'}), 403
    
    data = request.json
    
    try:
        # Update .env file
        env_file = '.env'
        env_content = {}
        
        # Read existing
        if os.path.exists(env_file):
            with open(env_file, 'r') as f:
                for line in f:
                    if '=' in line and not line.startswith('#'):
                        key, value = line.strip().split('=', 1)
                        env_content[key] = value
        
        # Update with new values
        if 'cognito_domain' in data:
            env_content['COGNITO_DOMAIN'] = data['cognito_domain']
        if 'cognito_client_id' in data:
            env_content['COGNITO_CLIENT_ID'] = data['cognito_client_id']
        if 'cognito_client_secret' in data:
            env_content['COGNITO_CLIENT_SECRET'] = data['cognito_client_secret']
        if 'proctor_emails' in data:
            env_content['PROCTOR_EMAILS'] = data['proctor_emails']
        if 'encryption_key' in data:
            env_content['ENCRYPTION_KEY'] = data['encryption_key']
        
        # Write back
        with open(env_file, 'w') as f:
            for key, value in env_content.items():
                f.write(f"{key}={value}\n")
        
        logger.info(f"Settings updated by {user_email}")
        
        return jsonify({
            'status': 'success',
            'message': 'Settings saved. Flask will need to restart for changes to take effect.'
        })
    
    except Exception as e:
        logger.error(f"Error updating settings: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/system/status', methods=['GET'])
@login_required
def system_status():
    """Get system status."""
    user_email = session.get('user_email', '')
    if not _is_proctor(user_email):
        return jsonify({'error': 'Access denied'}), 403
    
    import psutil
    import subprocess
    
    try:
        # Get CPU and memory
        cpu_percent = psutil.cpu_percent(interval=1)
        memory = psutil.virtual_memory()
        
        # Get Flask service status
        result = subprocess.run(
            ['/usr/bin/systemctl', 'is-active', 'flask-app'],
            capture_output=True,
            text=True,
            timeout=5
        )
        flask_status = result.stdout.strip()
        
        # Get last restart time
        result = subprocess.run(
            ['/usr/bin/systemctl', 'show', 'flask-app', '-p', 'StateChangeTimestamp'],
            capture_output=True,
            text=True,
            timeout=5
        )
        restart_time = result.stdout.strip()
        
        return jsonify({
            'cpu_percent': cpu_percent,
            'memory_percent': memory.percent,
            'memory_used_mb': memory.used // (1024 ** 2),
            'memory_total_mb': memory.total // (1024 ** 2),
            'flask_status': flask_status,
            'restart_time': restart_time
        })
    
    except Exception as e:
        logger.warning(f"Could not get system status: {str(e)}")
        return jsonify({
            'error': 'Could not retrieve system status',
            'details': str(e)
        }), 500


@app.route('/api/admin/system/restart', methods=['POST'])
@login_required
def restart_flask():
    """Restart Flask application."""
    user_email = session.get('user_email', '')
    if not _is_proctor(user_email):
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        import subprocess
        
        logger.info(f"Flask restart requested by {user_email}")
        
        subprocess.run(
            ['/usr/bin/sudo', '/usr/bin/systemctl', 'restart', 'flask-app'],
            timeout=10
        )
        
        return jsonify({
            'status': 'success',
            'message': 'Flask is restarting. Page will reload in 5 seconds.'
        })
    
    except Exception as e:
        logger.error(f"Error restarting Flask: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/system/logs', methods=['GET'])
@login_required
def get_logs():
    """Get recent Flask logs, with optional filters:
      - user: case-insensitive substring match on email (or any text)
      - q: case-insensitive substring match on free text
      - level: ERROR | WARNING | INFO - matches the level Flask's logger
        writes into each line (e.g. "app - ERROR - ...")
      - since: minutes; only journal lines from the last N minutes

    All filters are ANDed together. Filtering pulls a much bigger tail
    window since matches could otherwise easily fall outside the default
    last-100-lines view.
    """
    user_email = session.get('user_email', '')
    if not _is_proctor(user_email):
        return jsonify({'error': 'Access denied'}), 403

    user_filter = (request.args.get('user') or '').strip()
    text_filter = (request.args.get('q') or '').strip()
    level_filter = (request.args.get('level') or '').strip().upper()
    since_minutes = (request.args.get('since') or '').strip()

    any_filter = bool(user_filter or text_filter or level_filter or since_minutes)
    tail_lines = '5000' if any_filter else '100'

    try:
        import subprocess

        cmd = ['/usr/bin/sudo', '/usr/bin/journalctl', '-u', 'flask-app', '--no-pager']
        if since_minutes.isdigit():
            cmd += ['--since', f'{since_minutes} minutes ago']
        else:
            cmd += ['-n', tail_lines]

        result = subprocess.run(cmd, capture_output=True, text=True, timeout=5)

        # journalctl's output always ends with a trailing newline, which
        # would become a leading blank line once we reverse the order below.
        logs = result.stdout.split('\n')
        while logs and logs[-1] == '':
            logs.pop()

        if user_filter:
            needle = user_filter.lower()
            logs = [line for line in logs if needle in line.lower()]
        if text_filter:
            needle = text_filter.lower()
            logs = [line for line in logs if needle in line.lower()]
        if level_filter in ('ERROR', 'WARNING', 'INFO'):
            # Flask's default formatter writes "LEVEL" as its own token
            # (e.g. "... - ERROR - ..."), so match on word boundaries to
            # avoid "ERROR" also matching inside unrelated words.
            import re
            pattern = re.compile(r'\b' + re.escape(level_filter) + r'\b')
            logs = [line for line in logs if pattern.search(line)]

        if any_filter:
            # Keep the response light even if filters match a lot.
            logs = logs[-300:]

        # journalctl returns oldest-first; flip so the newest entry is
        # always at the top of the list (and therefore the top of the page).
        logs = list(reversed(logs))

        return jsonify({
            'status': 'success',
            'logs': logs,
            'filters': {
                'user': user_filter or None,
                'q': text_filter or None,
                'level': level_filter or None,
                'since_minutes': since_minutes or None
            }
        })
    
    except Exception as e:
        logger.warning(f"Could not retrieve logs: {str(e)}")
        return jsonify({
            'logs': ['Could not retrieve logs. Make sure you have sudo access.']
        })


@app.route('/api/admin/deploy/git-pull', methods=['POST'])
@login_required
def git_pull():
    """Pull latest code from git."""
    user_email = session.get('user_email', '')
    if not _is_proctor(user_email):
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        import subprocess
        
        logger.info(f"Git pull requested by {user_email}")
        
        result = subprocess.run(
            ['/usr/bin/git', 'pull', 'origin', 'main'],
            capture_output=True,
            text=True,
            timeout=30,
            cwd='/home/ubuntu/ai-assurance-lab'
        )
        
        # Restart Flask
        subprocess.run(
            ['/usr/bin/sudo', '/usr/bin/systemctl', 'restart', 'flask-app'],
            timeout=10
        )
        
        return jsonify({
            'status': 'success',
            'message': 'Code updated and Flask restarted',
            'output': result.stdout
        })
    
    except Exception as e:
        logger.error(f"Error pulling code: {str(e)}")
        return jsonify({'error': str(e)}), 500

# ============================================================================
# Static Pages
# ============================================================================

@app.route('/')
def index():
    """Home page - redirect to login if not authenticated."""
    if 'user_email' in session:
        return redirect(url_for('lab'))
    return redirect(url_for('login'))


# ============================================================================
# Error Handlers
# ============================================================================

@app.errorhandler(404)
def not_found(e):
    """Handle 404 errors."""
    logger.warning(f"404 error: {request.path}")
    return jsonify({'error': 'Not found'}), 404


@app.errorhandler(500)
def server_error(e):
    """Handle 500 errors."""
    logger.error(f"500 error: {str(e)}")
    return jsonify({'error': 'Internal server error'}), 500


if __name__ == '__main__':
    # Check for required environment variables
    required_vars = [
        'COGNITO_DOMAIN',
        'COGNITO_CLIENT_ID',
        'COGNITO_CLIENT_SECRET',
        'ENCRYPTION_KEY'
    ]
    
    missing = [var for var in required_vars if not os.getenv(var)]
    if missing:
        logger.error(f"Missing required environment variables: {', '.join(missing)}")
        logger.error("Please set these in .env file")
    
    logger.info(f"Starting Flask app on {APP_URL}")
    app.run(debug=os.getenv('FLASK_ENV') == 'development', host='0.0.0.0', port=5000)
